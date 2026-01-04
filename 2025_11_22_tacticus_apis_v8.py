#!/usr/bin/env python
# coding: utf-8

# In[105]:


import re
import requests
import pandas as pd
import numpy as np
import datetime as dt 
from time import gmtime, strftime
import warnings
import dropbox
from dropbox.oauth import DropboxOAuth2FlowNoRedirect
from dropbox.files import WriteMode
import openpyxl
from io import BytesIO
import os
#allow all columns to be visible
pd.set_option('display.max_columns', None)
#supress scientific notation
pd.set_option('display.float_format', lambda x: '%.9f' % x)
#color coding
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule


# In[107]:


#environmental variables for secrets use

dropbox_app_key = os.environ["APP_KEY"]
dropbox_app_secret = os.environ["APP_SECRET"]
dropbox_refresh_token = os.environ["REFRESH_TOKEN"]
api_bi = os.environ["api_bi"]
api_us = os.environ["api_us"]
api_vn = os.environ["api_vn"]
api_ky = os.environ["api_ky"]
dropbox_path = os.environ["dropbox_path"]


# In[108]:


"""
#previously generated mapping of all the guild members

global_member_list = pd.concat([
    bi_members[['userId','user_nicknames']],
    us_members[['userId','user_nicknames']],
    vn_members[['userId','user_nicknames']],
    ky_members[['userId','user_nicknames']]
    ], 
    axis=0, ignore_index=True)

with pd.ExcelWriter('global_member_list' + '.xlsx') as writer:
    global_member_list.to_excel(writer, sheet_name='global_member_list', index=False)

#persisted, to be then read from dropbox app folder
"""
# Create Dropbox client using refresh token
dbx = dropbox.Dropbox(
    oauth2_refresh_token = dropbox_refresh_token,
    app_key = dropbox_app_key,
    app_secret = dropbox_app_secret
)

metadata, res = dbx.files_download(dropbox_path)
file_content = res.content
global_member_list = pd.read_excel(BytesIO(file_content), engine='openpyxl')


# In[109]:


url_guild = 'https://api.tacticusgame.com/api/v1/guild'
url_raid_generic = 'https://api.tacticusgame.com/api/v1/guildRaid'


# In[110]:


#get raid season number programmatically and make raid link season-specific

headers = {"accept": "application/json", "X-API-KEY": api_us}
r_raid = requests.get(url_raid_generic, headers = headers)
data_raid_generic = r_raid.json()

#check whether there are any battles in the latest season log. If there are any - proceed. If everything is empty - pull data for previous season
if data_raid_generic['entries'] == []:
    raid_season = data_raid_generic['season'] - 1
else:
    raid_season = data_raid_generic['season']

# In[111]:


#read existing toplines file, to check for the latest season from api vs latest one in the existing file

dropbox_path_existing_toplines = "/global_toplines.xlsx"  # change to your path

metadata_toplines, res_toplines = dbx.files_download(dropbox_path_existing_toplines)
file_content_toplines = res_toplines.content
global_existing_toplines = pd.read_excel(BytesIO(file_content_toplines), engine='openpyxl')


# In[112]:


#if season from api is new - copy existing file into Archive and overwrite existing file

if raid_season == global_existing_toplines[['raid_season']].max().iloc[0]:
    print("same_season")
else:
    SRC_PATH = "/" +'global_toplines' + '.xlsx'
    DEST_PATH = "/" + "Archive/" + 'global_toplines_' + str(global_existing_toplines[['raid_season']].max().iloc[0]) + '.xlsx'

    # Download the original file
    _, res = dbx.files_download(SRC_PATH)
    file_content = res.content

    # Upload to the new location with overwrite
    dbx.files_upload(
        file_content,
        DEST_PATH,
        mode=WriteMode("overwrite")
    )
    
    print("added_into_archive")


# In[113]:


#define_main_input_function
def get_guild_data(guild_api, global_member_list, raid_season_input):
    #define season
    url_raid = 'https://api.tacticusgame.com/api/v1/guildRaid/' + str(raid_season_input)
    #pull guild data
    headers = {"accept": "application/json", "X-API-KEY": guild_api}
    r_guild = requests.get(url_guild, headers = headers)
    r_raid = requests.get(url_raid, headers = headers)

    data_guild = r_guild.json()
    data_raid = r_raid.json()
    
    guild_name = data_guild['guild']['name']
    guild_tag = data_guild['guild']['guildTag']
    guild_level = data_guild['guild']['level']
    
    #assemble data frame with members
    members = pd.DataFrame(data_guild['guild']['members'], columns=['userId', 'role', 'level', 'lastActivityOn'])
    df_members = pd.DataFrame(data_guild['guild']['members'], columns=['userId', 'role', 'level', 'lastActivityOn'])
    df_members = df_members.merge(global_member_list, on='userId', how='left')
    
    #assemble guild raid logs
    df_raid_log = pd.DataFrame(data_raid['entries'], columns=[
      'userId',
      'tier',
      'set',
      'encounterIndex',
      'remainingHp',
      'maxHp',
      'encounterType',
      'unitId',
      'type',
      'rarity',
      'damageDealt',
      'damageType',
      'startedOn',
      'completedOn',
      'heroDetails',
      'machineOfWarDetails',
      'globalConfigHash'
    ])
    
    #join data from guild table
    df_raid_log = df_raid_log.merge(global_member_list, on='userId', how='left')

    #get user level from members table
    df_raid_log = df_raid_log.merge(df_members[['userId', 'level']], on='userId', how='left')

    #add unique index number for each attack
    df_raid_log['attack_index'] = range(1, len(df_raid_log) + 1)

    #create dummy variables for units and MOWs
    df_raid_log['MOW_name'] = None
    df_raid_log['MOW_power'] = None

    df_raid_log['Unit_1_name'] = None
    df_raid_log['Unit_1_power'] = None

    df_raid_log['Unit_2_name'] = None
    df_raid_log['Unit_2_power'] = None

    df_raid_log['Unit_3_name'] = None
    df_raid_log['Unit_3_power'] = None

    df_raid_log['Unit_4_name'] = None
    df_raid_log['Unit_4_power'] = None

    df_raid_log['Unit_5_name'] = None
    df_raid_log['Unit_5_power'] = None
    
    #loop though units and MOWs fields and split them into separate columns
    for i in range(0,len(df_raid_log)-1):
        if df_raid_log.loc[i, 'damageType'] == "Bomb":
            continue
        elif df_raid_log.loc[i, 'damageType'] == "Battle":
            try:
                df_raid_log.loc[i, 'MOW_name'] = df_raid_log.loc[i, 'machineOfWarDetails']['unitId']
            except:
                df_raid_log.loc[i, 'MOW_name'] = np.nan  
            try:
                df_raid_log.loc[i, 'MOW_power'] = df_raid_log.loc[i, 'machineOfWarDetails']['power']
            except:
                df_raid_log.loc[i, 'MOW_power'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_1_name'] = df_raid_log.loc[i, 'heroDetails'][0]['unitId']
            except:
                df_raid_log.loc[i, 'Unit_1_name'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_1_power'] = df_raid_log.loc[i, 'heroDetails'][0]['power']
            except:
                df_raid_log.loc[i, 'Unit_1_power'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_2_name'] = df_raid_log.loc[i, 'heroDetails'][1]['unitId']
            except:
                df_raid_log.loc[i, 'Unit_2_name'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_2_power'] = df_raid_log.loc[i, 'heroDetails'][1]['power']
            except:
                df_raid_log.loc[i, 'Unit_2_power'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_3_name'] = df_raid_log.loc[i, 'heroDetails'][2]['unitId']
            except:
                df_raid_log.loc[i, 'Unit_3_name'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_3_power'] = df_raid_log.loc[i, 'heroDetails'][2]['power']
            except:
                df_raid_log.loc[i, 'Unit_3_power'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_4_name'] = df_raid_log.loc[i, 'heroDetails'][3]['unitId']
            except:
                df_raid_log.loc[i, 'Unit_4_name'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_4_power'] = df_raid_log.loc[i, 'heroDetails'][3]['power']
            except:
                df_raid_log.loc[i, 'Unit_4_power'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_5_name'] = df_raid_log.loc[i, 'heroDetails'][4]['unitId']
            except:
                df_raid_log.loc[i, 'Unit_5_name'] = np.nan
            try:
                df_raid_log.loc[i, 'Unit_5_power'] = df_raid_log.loc[i, 'heroDetails'][4]['power']
            except:
                df_raid_log.loc[i, 'Unit_5_power'] = np.nan
        else:
            continue
            
    #add flagging for meta teams
    df_raid_log['meta_mech_flag'] = np.where(df_raid_log['Unit_1_name'] == "admecRuststalker", 1,
                                    np.where(df_raid_log['Unit_2_name'] == "admecRuststalker", 1,
                                    np.where(df_raid_log['Unit_3_name'] == "admecRuststalker", 1,
                                    np.where(df_raid_log['Unit_4_name'] == "admecRuststalker", 1,
                                    np.where(df_raid_log['Unit_5_name'] == "admecRuststalker", 1,0)))))

    #Ragnar but not Karian
    df_raid_log['meta_multi_flag'] = np.where(
                                        (df_raid_log['Unit_1_name'] == "custoBladeChampion") |
                                        (df_raid_log['Unit_2_name'] == "custoBladeChampion") |
                                        (df_raid_log['Unit_3_name'] == "custoBladeChampion") |
                                        (df_raid_log['Unit_4_name'] == "custoBladeChampion") |
                                        (df_raid_log['Unit_5_name'] == "custoBladeChampion"), 0,
                                     np.where(df_raid_log['Unit_1_name'] == "spaceBlackmane", 1,
                                     np.where(df_raid_log['Unit_2_name'] == "spaceBlackmane", 1,
                                     np.where(df_raid_log['Unit_3_name'] == "spaceBlackmane", 1,
                                     np.where(df_raid_log['Unit_4_name'] == "spaceBlackmane", 1,
                                     np.where(df_raid_log['Unit_5_name'] == "spaceBlackmane", 1,0))))))

    df_raid_log['meta_neuro_flag'] = np.where(df_raid_log['Unit_1_name'] == "tyranNeurothrope", 1,
                                     np.where(df_raid_log['Unit_2_name'] == "tyranNeurothrope", 1,
                                     np.where(df_raid_log['Unit_3_name'] == "tyranNeurothrope", 1,
                                     np.where(df_raid_log['Unit_4_name'] == "tyranNeurothrope", 1,
                                     np.where(df_raid_log['Unit_5_name'] == "tyranNeurothrope", 1,0)))))

    df_raid_log['meta_custodes_flag'] = np.where(df_raid_log['Unit_1_name'] == "custoBladeChampion", 1,
                                        np.where(df_raid_log['Unit_2_name'] == "custoBladeChampion", 1,
                                        np.where(df_raid_log['Unit_3_name'] == "custoBladeChampion", 1,
                                        np.where(df_raid_log['Unit_4_name'] == "custoBladeChampion", 1,
                                        np.where(df_raid_log['Unit_5_name'] == "custoBladeChampion", 1,0)))))

    #remove cases where damage is 0 from raid logs
    df_raid_log = df_raid_log.loc[df_raid_log['damageDealt'] > 0]

    #add guild name
    df_raid_log['guild'] = guild_name

    #new component - finishing battle flag
    df_raid_log['finishing_battle_flag'] = np.where((df_raid_log['remainingHp'] == 0) & 
                                                    (df_raid_log['damageType'] == 'Battle') &
                                                    (df_raid_log['damageDealt'] < df_raid_log['maxHp']), 
                                                    1, 0)

    #defining the legitimate attack - should be on the enemy with at least 6x hp vs the average bomb in the guild
    avg_bomb_damage_per_guild = int(np.average(df_raid_log.loc[df_raid_log['damageType'] == 'Bomb', 'damageDealt']))

    #flag whether the finishing battle is legitimate
    finishing_multiplier = 6
    df_raid_log['legitimate_finishing_battle_flag'] = np.where((df_raid_log['finishing_battle_flag'] == 1) & 
                                                               (df_raid_log['damageDealt'] >= avg_bomb_damage_per_guild * finishing_multiplier),
                                                               1, 0)
            
    #aggregated raid data on player level
    aggregated_raid_data = (
        df_raid_log
        .groupby(['user_nicknames'])
        .apply(lambda g: pd.Series({
            #level
            'user_level': g['level'].max(),
            
            #num attacks
            'num_attacks': g['attack_index'].count(),
            'num_bombs': (g['damageType'] == 'Bomb').sum(),
            'num_battles': (g['damageType'] == 'Battle').sum(),

            'num_attacks_bosses': g.loc[g['encounterType'] == 'Boss', 'attack_index'].count(),
            'num_bombs_bosses': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'Boss'),'attack_index'].count(),
            'num_battles_bosses': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'Boss'),'attack_index'].count(),

            'num_attacks_side_bosses': g.loc[g['encounterType'] == 'SideBoss', 'attack_index'].count(),
            'num_bombs_side_bosses': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'SideBoss'),'attack_index'].count(),
            'num_battles_side_bosses': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'SideBoss'),'attack_index'].count(),

            #damage
            'damage_attacks': g['damageDealt'].sum(),
            'damage_bombs': g.loc[g['damageType'] == 'Bomb', 'damageDealt'].sum(),
            'damage_battles': g.loc[g['damageType'] == 'Battle', 'damageDealt'].sum(),

            'damage_attacks_bosses': g.loc[g['encounterType'] == 'Boss', 'damageDealt'].sum(),
            'damage_bombs_bosses': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'Boss'),'damageDealt'].sum(),
            'damage_battles_bosses': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'Boss'),'damageDealt'].sum(),

            'damage_attacks_side_bosses': g.loc[g['encounterType'] == 'SideBoss', 'damageDealt'].sum(),
            'damage_bombs_side_bosses': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'SideBoss'),'damageDealt'].sum(),
            'damage_battles_side_bosses': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'SideBoss'),'damageDealt'].sum(),

            #legendaries num attacks - tier 4 is first circle of legendaries
            'num_attacks_legendary': g.loc[g['tier'] >= 4, 'attack_index'].count(),
            'num_bombs_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['tier'] >= 4),'attack_index'].count(),
            'num_battles_legendary': g.loc[(g['damageType'] == 'Battle') & (g['tier'] >= 4),'attack_index'].count(),

            'num_attacks_bosses_legendary': g.loc[(g['encounterType'] == 'Boss') & (g['tier'] >= 4), 'attack_index'].count(),
            'num_bombs_bosses_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'Boss') & (g['tier'] >= 4),'attack_index'].count(),
            'num_battles_bosses_legendary': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'Boss') & (g['tier'] >= 4),'attack_index'].count(),

            'num_attacks_side_bosses_legendary': g.loc[(g['encounterType'] == 'SideBoss') & (g['tier'] >= 4), 'attack_index'].count(),
            'num_bombs_side_bosses_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'SideBoss') & (g['tier'] >= 4),'attack_index'].count(),
            'num_battles_side_bosses_legendary': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'SideBoss') & (g['tier'] >= 4),'attack_index'].count(),

            #legendaries damage
            'damage_attacks_legendary': g.loc[g['tier'] >= 4, 'damageDealt'].sum(),
            'damage_bombs_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['tier'] >= 4),'damageDealt'].sum(),
            'damage_battles_legendary': g.loc[(g['damageType'] == 'Battle') & (g['tier'] >= 4),'damageDealt'].sum(),

            'damage_attacks_bosses_legendary': g.loc[(g['encounterType'] == 'Boss') & (g['tier'] >= 4), 'damageDealt'].sum(),
            'damage_bombs_bosses_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'Boss') & (g['tier'] >= 4),'damageDealt'].sum(),
            'damage_battles_bosses_legendary': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'Boss') & (g['tier'] >= 4),'damageDealt'].sum(),

            'damage_attacks_side_bosses_legendary': g.loc[(g['encounterType'] == 'SideBoss') & (g['tier'] >= 4), 'damageDealt'].sum(),
            'damage_bombs_side_bosses_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'SideBoss') & (g['tier'] >= 4),'damageDealt'].sum(),
            'damage_battles_side_bosses_legendary': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'SideBoss') & (g['tier'] >= 4),'damageDealt'].sum(),

            #meta teams flagging
            'num_battles_meta_mech': g.loc[(g['damageType'] == 'Battle') & (g['meta_mech_flag'] == 1),'attack_index'].count(),
            'num_battles_meta_multi': g.loc[(g['damageType'] == 'Battle') & (g['meta_multi_flag'] == 1),'attack_index'].count(),
            'num_battles_meta_neuro': g.loc[(g['damageType'] == 'Battle') & (g['meta_neuro_flag'] == 1),'attack_index'].count(),
            'num_battles_meta_custodes': g.loc[(g['damageType'] == 'Battle') & (g['meta_custodes_flag'] == 1),'attack_index'].count(),
            
            'damage_meta_mech': g.loc[(g['damageType'] == 'Battle') & (g['meta_mech_flag'] == 1),'damageDealt'].sum(),
            'damage_meta_multi': g.loc[(g['damageType'] == 'Battle') & (g['meta_multi_flag'] == 1),'damageDealt'].sum(),
            'damage_meta_neuro': g.loc[(g['damageType'] == 'Battle') & (g['meta_neuro_flag'] == 1),'damageDealt'].sum(),
            'damage_meta_custodes': g.loc[(g['damageType'] == 'Battle') & (g['meta_custodes_flag'] == 1),'damageDealt'].sum(),
            
            #meta teams with legendaries
            'num_battles_meta_mech_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_mech_flag'] == 1) & (g['tier'] >= 4),'attack_index'].count(),
            'num_battles_meta_multi_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_multi_flag'] == 1) & (g['tier'] >= 4),'attack_index'].count(),
            'num_battles_meta_neuro_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_neuro_flag'] == 1) & (g['tier'] >= 4),'attack_index'].count(),
            'num_battles_meta_custodes_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_custodes_flag'] == 1) & (g['tier'] >= 4),'attack_index'].count(),
            
            'damage_meta_mech_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_mech_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].sum(),
            'damage_meta_multi_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_multi_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].sum(),
            'damage_meta_neuro_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_neuro_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].sum(),
            'damage_meta_custodes_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_custodes_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].sum(),
            
            #averages
            'avg_damage_attacks': g['damageDealt'].mean(),
            'avg_damage_bombs': g.loc[g['damageType'] == 'Bomb', 'damageDealt'].mean(),
            'avg_damage_battles': g.loc[g['damageType'] == 'Battle', 'damageDealt'].mean(),

            'avg_damage_attacks_bosses': g.loc[g['encounterType'] == 'Boss', 'damageDealt'].mean(),
            'avg_damage_bombs_bosses': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'Boss'),'damageDealt'].mean(),
            'avg_damage_battles_bosses': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'Boss'),'damageDealt'].mean(),

            'avg_damage_attacks_side_bosses': g.loc[g['encounterType'] == 'SideBoss', 'damageDealt'].mean(),
            'avg_damage_bombs_side_bosses': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'SideBoss'),'damageDealt'].mean(),
            'avg_damage_battles_side_bosses': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'SideBoss'),'damageDealt'].mean(),

            'avg_damage_attacks_legendary': g.loc[g['tier'] >= 4, 'damageDealt'].mean(),
            'avg_damage_bombs_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['tier'] >= 4),'damageDealt'].mean(),
            'avg_damage_battles_legendary': g.loc[(g['damageType'] == 'Battle') & (g['tier'] >= 4),'damageDealt'].mean(),

            'avg_damage_attacks_bosses_legendary': g.loc[(g['encounterType'] == 'Boss') & (g['tier'] >= 4), 'damageDealt'].mean(),
            'avg_damage_bombs_bosses_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'Boss') & (g['tier'] >= 4),'damageDealt'].mean(),
            'avg_damage_battles_bosses_legendary': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'Boss') & (g['tier'] >= 4),'damageDealt'].mean(),

            'avg_damage_attacks_side_bosses_legendary': g.loc[(g['encounterType'] == 'SideBoss') & (g['tier'] >= 4), 'damageDealt'].mean(),
            'avg_damage_bombs_side_bosses_legendary': g.loc[(g['damageType'] == 'Bomb') & (g['encounterType'] == 'SideBoss') & (g['tier'] >= 4),'damageDealt'].mean(),
            'avg_damage_battles_side_bosses_legendary': g.loc[(g['damageType'] == 'Battle') & (g['encounterType'] == 'SideBoss') & (g['tier'] >= 4),'damageDealt'].mean(),
            
            'avg_damage_meta_mech': g.loc[(g['damageType'] == 'Battle') & (g['meta_mech_flag'] == 1),'damageDealt'].mean(),
            'avg_damage_meta_multi': g.loc[(g['damageType'] == 'Battle') & (g['meta_multi_flag'] == 1),'damageDealt'].mean(),
            'avg_damage_meta_neuro': g.loc[(g['damageType'] == 'Battle') & (g['meta_neuro_flag'] == 1),'damageDealt'].mean(),
            'avg_damage_meta_custodes': g.loc[(g['damageType'] == 'Battle') & (g['meta_custodes_flag'] == 1),'damageDealt'].mean(),
                        
            'avg_damage_meta_mech_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_mech_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].mean(),
            'avg_damage_meta_multi_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_multi_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].mean(),
            'avg_damage_meta_neuro_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_neuro_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].mean(),
            'avg_damage_meta_custodes_legendary': g.loc[(g['damageType'] == 'Battle') & (g['meta_custodes_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].mean(),

            #legendary finishing attacks and legitimate finishing attacks
            'num_legendary_finishing_battles': g.loc[(g['finishing_battle_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].count(),
            'num_legendary_legitimate_finishing_battles': g.loc[(g['legitimate_finishing_battle_flag'] == 1) & (g['tier'] >= 4),'damageDealt'].count(),
            'avg_damage_battles_legendary_non_finishing': g.loc[(g['damageType'] == 'Battle') & (g['tier'] >= 4) & (g['legitimate_finishing_battle_flag'] == 0),'damageDealt'].mean()
        }))
        .reset_index()
    )
    
    #remove decimals before calculating efficiency and points
    aggregated_raid_data = aggregated_raid_data.round()
    
    #replace NaNs with zeroes
    aggregated_raid_data = aggregated_raid_data.fillna(0)
    
    #add guild name as first column
    aggregated_raid_data.insert(0, 'guild', guild_name)  

    ################################################calculate toplines per bosses and side bosses
    #catch formatted version of side bosses and bosses names
    pattern = r'(?:.*?\d+){2}(.*)$'
    df_raid_log['unit_name'] = df_raid_log['unitId'].str.extract(pattern)
    df_raid_log['unit_name'] = df_raid_log['rarity'] + df_raid_log['encounterType'] + "_" + df_raid_log['set'].values.astype(str) + "_" + df_raid_log['unit_name']

    #create boss df, split it into regular and legitimate finishing attacks, then merge them to create separate columns for KPIs for both
    #regular attacks
    boss_df_regular = (
        df_raid_log[
            (df_raid_log['damageType'] == 'Battle') &
            (df_raid_log['tier'] >= 4) &
            (df_raid_log['legitimate_finishing_battle_flag'] == 0)
        ]
        .groupby(['user_nicknames', 'unit_name'])['damageDealt']
        .agg(
            num_battles='count',
            total_damage='sum',
            avg_damage='mean'
        )
        .reset_index()
    )

    #add guild name to be used further on
    boss_df_regular['guild'] = guild_name
    boss_df_regular = boss_df_regular[['guild','user_nicknames','unit_name','num_battles','total_damage','avg_damage']]

    #legit finishing attacks
    boss_df_legit_finish = (
        df_raid_log[
            (df_raid_log['damageType'] == 'Battle') &
            (df_raid_log['tier'] >= 4) &
            (df_raid_log['legitimate_finishing_battle_flag'] == 1)
        ]
        .groupby(['user_nicknames', 'unit_name'])['damageDealt']
        .agg(
            num_finish_battles='count',
            total_finish_damage='sum',
            avg_finish_damage='mean'
        )
        .reset_index()
    )

    #add guild name to be used further on
    boss_df_legit_finish['guild'] = guild_name
    boss_df_legit_finish = boss_df_legit_finish[['guild','user_nicknames','unit_name',
                                                 'num_finish_battles','total_finish_damage','avg_finish_damage']]

    boss_df = boss_df_regular.merge(boss_df_legit_finish, on = ['guild','user_nicknames','unit_name'], how = 'outer')
    boss_df = boss_df.fillna(0)
    boss_df = boss_df.round()

    #pivot and format toplines per bosses
    pivot_boss_df = boss_df.pivot(index='user_nicknames', columns='unit_name', 
                                  values=['num_battles', 'total_damage', 'avg_damage', 
                                         'num_finish_battles','total_finish_damage','avg_finish_damage'])
    pivot_boss_df.columns = ['_'.join(str(s).strip() for s in col if s) for col in pivot_boss_df.columns]
    pivot_boss_df.reset_index(inplace=True)
    pivot_boss_df = pivot_boss_df.fillna(0)
    pivot_boss_df = pivot_boss_df.round()

    #merge with main aggr data
    aggregated_raid_data = aggregated_raid_data.merge(pivot_boss_df, on='user_nicknames', how='left')

    #calculate number of additional circles per guild, add it to aggr frame
    add_circles = (max(df_raid_log['tier']) - 5)/2
    aggregated_raid_data['add_circles'] = add_circles

    #add flag for which meta team is dominant
    aggregated_raid_data['max_archetype'] = aggregated_raid_data[[
        'num_battles_meta_mech', 
        'num_battles_meta_multi', 
        'num_battles_meta_neuro', 
        'num_battles_meta_custodes'
    ]].idxmax(axis=1)

    aggregated_raid_data['max_archetype'] = aggregated_raid_data['max_archetype'].map({
        'num_battles_meta_mech': 'mech',
        'num_battles_meta_multi': 'multi',
        'num_battles_meta_neuro': 'neuro',
        'num_battles_meta_custodes': 'custodes',
    })


    return df_members, df_raid_log, aggregated_raid_data, boss_df


# In[114]:


#set dummy dfs to be further populated
bi_members = pd.DataFrame()
us_members = pd.DataFrame()
vn_members = pd.DataFrame()
ky_members = pd.DataFrame()

bi_source_raid_log = pd.DataFrame()
us_source_raid_log = pd.DataFrame()
vn_source_raid_log = pd.DataFrame()
ky_source_raid_log = pd.DataFrame()

bi_boss_df = pd.DataFrame()
us_boss_df = pd.DataFrame()
vn_boss_df = pd.DataFrame()
ky_boss_df = pd.DataFrame()

us_aggr_raid_log = pd.DataFrame() 
bi_aggr_raid_log = pd.DataFrame()
vn_aggr_raid_log = pd.DataFrame()
ky_aggr_raid_log = pd.DataFrame()

#run the function to pull and format data
try:
    bi_members, bi_source_raid_log, bi_aggr_raid_log, bi_boss_df = get_guild_data(api_bi, global_member_list, raid_season)
    print("bi_done")
except Exception:
    print("bi_error")
    
try:
    us_members, us_source_raid_log, us_aggr_raid_log, us_boss_df = get_guild_data(api_us, global_member_list, raid_season)
    print("us_done")
except Exception:
    print("us_error")

try:
    vn_members, vn_source_raid_log, vn_aggr_raid_log, vn_boss_df = get_guild_data(api_vn, global_member_list, raid_season)
    print("vn_done")
except Exception:
    print("vn_error")
    

try:
    ky_members, ky_source_raid_log, ky_aggr_raid_log, ky_boss_df = get_guild_data(api_ky, global_member_list, raid_season)
    print("ky_done")
except Exception:
    print("ky_error")


# In[115]:


#create an empty dfs to be populated with further pulls
benchmark_total_boss_df = pd.DataFrame()
global_boss_df = pd.DataFrame()
global_aggr_raid_log = pd.DataFrame()

processed_logs = [
    bi_source_raid_log, 
    us_source_raid_log,
    vn_source_raid_log,
    ky_source_raid_log
]

processed_boss_logs = [
    bi_boss_df,
    us_boss_df,
    vn_boss_df,
    ky_boss_df
]

processed_aggr_logs = [
    us_aggr_raid_log, 
    bi_aggr_raid_log, 
    vn_aggr_raid_log, 
    ky_aggr_raid_log
]

#merge dfs in case they were properly generated, or skip them in case of error
for df_order in range(0,len(processed_logs)):
    benchmark_total_boss_df = pd.concat([benchmark_total_boss_df, processed_logs[df_order]], axis=0, ignore_index=True)
    global_boss_df = pd.concat([global_boss_df, processed_boss_logs[df_order]], axis=0, ignore_index=True)
    global_aggr_raid_log = pd.concat([global_aggr_raid_log, processed_aggr_logs[df_order]], axis=0, ignore_index=True)


# In[123]:


#create copy of the logs to export further on
export_full_logs = benchmark_total_boss_df
#remove ids for privacy reasons
export_full_logs = export_full_logs.drop('userId', axis=1)
#put couple of helpful columns in front
front_cols = ['guild', 'user_nicknames', 'attack_index']
export_full_logs = export_full_logs[front_cols + [c for c in export_full_logs.columns if c not in front_cols]]


# In[117]:


#calculate benchmarks for boss damage
#remove any finishing hits, regardless whether legitimate from benchmarking calculations
benchmark_total_boss_df = benchmark_total_boss_df.loc[
    (benchmark_total_boss_df['damageType'] == 'Battle') &
    (benchmark_total_boss_df['tier'] >= 4) &
    (benchmark_total_boss_df['finishing_battle_flag'] == 0)
].groupby(['guild', 'unit_name']).apply(lambda g: pd.Series({
    'benchmark_num_battles': g['damageDealt'].count(),
    'benchmark_total_damage': g['damageDealt'].sum(),
    'benchmark_avg_damage': g['damageDealt'].mean()
})).reset_index()

benchmark_total_boss_df = benchmark_total_boss_df.groupby(['unit_name']).apply(lambda g: pd.Series({
    'benchmark_max_avg_damage': g['benchmark_avg_damage'].max()
})).reset_index()


# In[118]:


#calculate efficiency for individual bosses for everybody
global_boss_df = global_boss_df.merge(benchmark_total_boss_df[['unit_name','benchmark_max_avg_damage']], on='unit_name', how='left')

global_boss_df['global_efficiency'] = global_boss_df['avg_damage'] / global_boss_df['benchmark_max_avg_damage']
global_boss_df['global_efficiency'] = global_boss_df['global_efficiency'].round(3)
global_boss_df = global_boss_df.fillna(0)

#take into account finishing hits, with the following logic:
#if there are both finishing and regular hits for the same boss - take efficiency from regular hits and * by num regulat + finishing hits for points
#if there are only finishing hits - take average efficiency per player on regular hits and use it instead

#find playerwise efficiency on non-finishing battles to be used as a plug
global_boss_df_playerwise_efficiency = global_boss_df.loc[
    (global_boss_df['num_battles'] > 0)
].groupby(['guild', 'user_nicknames']).apply(lambda g: pd.Series({
    'benchmark_avg_efficiency_plug': g['global_efficiency'].mean()
})).reset_index()

global_boss_df = global_boss_df.merge(global_boss_df_playerwise_efficiency, on=['guild', 'user_nicknames'], how='left')

global_boss_df['global_points'] = np.where((global_boss_df['num_battles'] > 0) & (global_boss_df['num_finish_battles'] > 0),
                                    global_boss_df['global_efficiency'] * (global_boss_df['num_battles'] + global_boss_df['num_finish_battles']),
                                  np.where((global_boss_df['num_battles'] > 0) & (global_boss_df['num_finish_battles'] == 0),
                                    global_boss_df['global_efficiency'] * global_boss_df['num_battles'],
                                  np.where((global_boss_df['num_battles'] == 0) & (global_boss_df['num_finish_battles'] > 0),
                                    global_boss_df['benchmark_avg_efficiency_plug'] * global_boss_df['num_finish_battles'],
                                    0)))

global_boss_df = global_boss_df.round(3)

global_boss_df['guild_and_name'] = global_boss_df['guild'] + global_boss_df['user_nicknames']


# In[119]:


#calculate total global efficiency, aggregated across all bosses on a player level
aggr_global_boss_df = global_boss_df.groupby(['guild','user_nicknames']).apply(lambda g: pd.Series({
    'total_points': g['global_points'].sum()
})).reset_index()

aggr_global_boss_df = aggr_global_boss_df.sort_values(by='total_points',ascending=False)

aggr_global_boss_df['guild_and_name'] = aggr_global_boss_df['guild'] + aggr_global_boss_df['user_nicknames']


# In[120]:


#calculate total points per boss, merge with aggr_global_boss_df
pivot_global_boss_df = global_boss_df.pivot(index='guild_and_name', columns='unit_name', values=['global_points'])
pivot_global_boss_df.columns = ['_'.join(str(s).strip() for s in col if s) for col in pivot_global_boss_df.columns]
pivot_global_boss_df.reset_index(inplace=True)
pivot_global_boss_df = pivot_global_boss_df.fillna(0)
pivot_global_boss_df = pivot_global_boss_df.round(3)

aggr_global_boss_df = aggr_global_boss_df.merge(pivot_global_boss_df, on=['guild_and_name'], how='left')

#create topline version of global export
global_detailed_toplines = global_aggr_raid_log[[
    'guild',
    'user_nicknames',
    'user_level',
    'max_archetype',
    "add_circles",
    "num_battles",
    "num_bombs",
    "avg_damage_battles",
    "num_battles_legendary",
    "avg_damage_battles_legendary",
    "num_legendary_finishing_battles",	
    "num_legendary_legitimate_finishing_battles",
    "avg_damage_battles_legendary_non_finishing",
    'num_battles_meta_mech_legendary',
    'num_battles_meta_custodes_legendary',
    'num_battles_meta_neuro_legendary',
    'num_battles_meta_multi_legendary'
]].merge(aggr_global_boss_df, on=['guild','user_nicknames'], how='left')

global_detailed_toplines = global_detailed_toplines.sort_values(by='total_points',ascending=False)

global_aggr_toplines = global_detailed_toplines[[
    "guild",
    "user_nicknames",
    'user_level',
    'max_archetype',
    "num_battles",
    "total_points"
]]

global_aggr_toplines['raid_season'] = raid_season

global_aggr_toplines = global_aggr_toplines[[
    'raid_season',
    "guild",
    "user_nicknames",
    'user_level',
    'max_archetype',
    "num_battles",
    "total_points"
]]

#remove redundant columns
global_detailed_toplines = global_detailed_toplines.drop('guild_and_name', axis=1)
global_boss_df = global_boss_df.drop('guild_and_name', axis=1)

######################## create new dataframe, with damage toplines for all bosses across meta teams

# In[new cell 1]:

# get data for last 5 seasons
bi_members_s0, bi_source_raid_log_s0, bi_aggr_raid_log_s0, bi_boss_df_s0 = get_guild_data(api_bi, global_member_list, raid_season)
us_members_s0, us_source_raid_log_s0, us_aggr_raid_log_s0, us_boss_df_s0 = get_guild_data(api_us, global_member_list, raid_season)
vn_members_s0, vn_source_raid_log_s0, vn_aggr_raid_log_s0, vn_boss_df_s0 = get_guild_data(api_vn, global_member_list, raid_season)
ky_members_s0, ky_source_raid_log_s0, ky_aggr_raid_log_s0, ky_boss_df_s0 = get_guild_data(api_ky, global_member_list, raid_season)

bi_members_s1, bi_source_raid_log_s1, bi_aggr_raid_log_s1, bi_boss_df_s1 = get_guild_data(api_bi, global_member_list, raid_season-1)
us_members_s1, us_source_raid_log_s1, us_aggr_raid_log_s1, us_boss_df_s1 = get_guild_data(api_us, global_member_list, raid_season-1)
vn_members_s1, vn_source_raid_log_s1, vn_aggr_raid_log_s1, vn_boss_df_s1 = get_guild_data(api_vn, global_member_list, raid_season-1)
ky_members_s1, ky_source_raid_log_s1, ky_aggr_raid_log_s1, ky_boss_df_s1 = get_guild_data(api_ky, global_member_list, raid_season-1)

bi_members_s2, bi_source_raid_log_s2, bi_aggr_raid_log_s2, bi_boss_df_s2 = get_guild_data(api_bi, global_member_list, raid_season-2)
us_members_s2, us_source_raid_log_s2, us_aggr_raid_log_s2, us_boss_df_s2 = get_guild_data(api_us, global_member_list, raid_season-2)
vn_members_s2, vn_source_raid_log_s2, vn_aggr_raid_log_s2, vn_boss_df_s2 = get_guild_data(api_vn, global_member_list, raid_season-2)
ky_members_s2, ky_source_raid_log_s2, ky_aggr_raid_log_s2, ky_boss_df_s2 = get_guild_data(api_ky, global_member_list, raid_season-2)

bi_members_s3, bi_source_raid_log_s3, bi_aggr_raid_log_s3, bi_boss_df_s3 = get_guild_data(api_bi, global_member_list, raid_season-3)
us_members_s3, us_source_raid_log_s3, us_aggr_raid_log_s3, us_boss_df_s3 = get_guild_data(api_us, global_member_list, raid_season-3)
vn_members_s3, vn_source_raid_log_s3, vn_aggr_raid_log_s3, vn_boss_df_s3 = get_guild_data(api_vn, global_member_list, raid_season-3)
ky_members_s3, ky_source_raid_log_s3, ky_aggr_raid_log_s3, ky_boss_df_s3 = get_guild_data(api_ky, global_member_list, raid_season-3)

bi_members_s4, bi_source_raid_log_s4, bi_aggr_raid_log_s4, bi_boss_df_s4 = get_guild_data(api_bi, global_member_list, raid_season-4)
us_members_s4, us_source_raid_log_s4, us_aggr_raid_log_s4, us_boss_df_s4 = get_guild_data(api_us, global_member_list, raid_season-4)
vn_members_s4, vn_source_raid_log_s4, vn_aggr_raid_log_s4, vn_boss_df_s4 = get_guild_data(api_vn, global_member_list, raid_season-4)
ky_members_s4, ky_source_raid_log_s4, ky_aggr_raid_log_s4, ky_boss_df_s4 = get_guild_data(api_ky, global_member_list, raid_season-4)


# concat dataframes across guilds for last 5 seasons
concat_raid_log = pd.concat([    
    bi_source_raid_log_s0, 
    us_source_raid_log_s0,
    vn_source_raid_log_s0,
    ky_source_raid_log_s0,
    
    bi_source_raid_log_s1, 
    us_source_raid_log_s1,
    vn_source_raid_log_s1,
    ky_source_raid_log_s1,

    bi_source_raid_log_s2, 
    us_source_raid_log_s2,
    vn_source_raid_log_s2,
    ky_source_raid_log_s2,

    bi_source_raid_log_s3, 
    us_source_raid_log_s3,
    vn_source_raid_log_s3,
    ky_source_raid_log_s3,

    bi_source_raid_log_s4, 
    us_source_raid_log_s4,
    vn_source_raid_log_s4,
    ky_source_raid_log_s4
], axis=0, ignore_index=True)

# find max damage per meta team
meta_boss_df = concat_raid_log.loc[
    (concat_raid_log['damageType'] == 'Battle') &
    (concat_raid_log['tier'] >= 4)
].groupby(['unit_name']).apply(lambda g: pd.Series({
    'set': g['set'].max(),
    'max_mech_damage': g.loc[(g['meta_mech_flag'] == 1),'damageDealt'].max(),
    'max_multi_damage': g.loc[(g['meta_multi_flag'] == 1),'damageDealt'].max(),
    'max_neuro_damage': g.loc[(g['meta_neuro_flag'] == 1),'damageDealt'].max(),
    'max_custodes_damage': g.loc[(g['meta_custodes_flag'] == 1),'damageDealt'].max(),
})).reset_index()

# format and sort the output
meta_boss_df  = meta_boss_df.fillna(0)
meta_boss_df_cols = meta_boss_df.columns[1:]
meta_boss_df[meta_boss_df_cols] = meta_boss_df[meta_boss_df_cols].apply(
    lambda c: pd.to_numeric(c, errors='coerce').round(0).astype('Int64')
)

meta_boss_df = meta_boss_df.sort_values(
    by=['set', 'unit_name'],
    ascending=[True, False]
)

meta_boss_df['_is_mythic'] = meta_boss_df['unit_name'].str.contains('Mythic', na=False)
meta_boss_df = pd.concat([    
    meta_boss_df.loc[meta_boss_df['_is_mythic'] == False],
    meta_boss_df.loc[meta_boss_df['_is_mythic'] == True],
])

meta_boss_df = meta_boss_df.drop(columns=['set'])
meta_boss_df = meta_boss_df.drop(columns=['_is_mythic'])

meta_boss_df[['rartiy_and_type', 'level', 'name']] = meta_boss_df['unit_name'].str.split('_', expand=True)

meta_boss_df = meta_boss_df[[
    "rartiy_and_type",
    "level",
    "name",
    "max_mech_damage",
    "max_multi_damage",
    "max_neuro_damage",
    "max_custodes_damage"
]]


# In[new_cell_4]:


######################### create a boss-wise data frame, with rows as ppl and pivoted stats per boss in columns
boss_wise_points = global_boss_df

boss_wise_points = boss_wise_points[[
    "guild",
    "user_nicknames",
    "unit_name",
    "num_battles",
    "avg_damage",
    "global_efficiency",
    "global_points"
]]

#get the unique list of bosses
unique_list = boss_wise_points['unit_name'].unique().tolist()

#merge total points and battles for further sorting
total_scored_df = global_aggr_toplines[[
    "guild",
    "user_nicknames",
    "num_battles",
    "total_points"
]].rename(columns={
    "num_battles": "overall_battles",
    "total_points": "overall_points"
})

boss_wise_points = boss_wise_points.merge(total_scored_df, on=['guild','user_nicknames'], how='left')

#sort the unique list of bosses
sorted_list = sorted(
    unique_list,
    key=lambda x: int(re.search(r'\d+', x).group())
)

sorted_list = sorted(
    sorted_list,
    key=lambda x: (
        'Mythic' in x,
        int(re.search(r'\d+', x).group())
    )
)

boss_wise_points_pivot = (
    boss_wise_points
    .pivot(
        index=['guild', 'user_nicknames', "overall_battles", "overall_points"],
        columns='unit_name',
        values=[
            'num_battles',
            'avg_damage',
            'global_efficiency',
            'global_points'
        ]
    )
    .swaplevel(0, 1, axis=1)
    .sort_index(axis=1, level=0, sort_remaining=False)
)

metric_order = [
    'num_battles',
    'avg_damage',
    'global_efficiency',
    'global_points'
]

unit_order = sorted_list

boss_wise_points_pivot = boss_wise_points_pivot.reindex(metric_order, axis=1, level=1)
boss_wise_points_pivot = boss_wise_points_pivot.reindex(unit_order, axis=1, level=0)

boss_wise_points_pivot.reset_index(inplace=True)
boss_wise_points_pivot = boss_wise_points_pivot.fillna(0)

boss_wise_points_pivot = boss_wise_points_pivot.sort_values(by='overall_points', ascending=False)


# In[circles_1]:
######################### create a circle-wise data frame, to check how many battles were spent for each boss each circle
def get_circle_data(raid_log_df):
    df1 = raid_log_df

    #extract boss name
    pattern = r'(?:.*?\d+){2}(.*)$'
    df1['unit_name'] = df1['unitId'].str.extract(pattern)

    #summarize damage per boss per circle
    df1 = df1.loc[
        (df1['damageType'] == 'Battle') &
        (df1['tier'] >= 4)
    ].groupby(['guild','unit_name', 'tier', 'set', 'rarity', 'encounterType']).apply(lambda g: pd.Series({
        'num_battles': g['damageDealt'].count(),
        'avg_damage': g['damageDealt'].mean()
    })).reset_index()

    #fix orders
    df1['set_order'] = np.where(df1['rarity'] == 'Mythic', df1['set']+5, df1['set'])
    df1['circles'] = np.ceil(((df1['tier']-3)/2)).astype(int)
    
    df1 = df1.sort_values(by=['set_order'], ascending=True)
    
    df1.drop(['tier'], axis=1, inplace=True)
    
    #pivot the df with bosses as rows and circles as columns
    df2 = (
        df1
        .pivot(
            index=['guild','unit_name','rarity','encounterType','set', 'set_order'],
            columns='circles',
            values=[
                'num_battles',
                'avg_damage'
            ]
        ).swaplevel(0, 1, axis=1)
        .sort_index(axis=1, level=0, sort_remaining=False)
    )
    
    
    df2 = df2.reset_index()
    df2 = df2.sort_values(by=['set_order','encounterType'], ascending = [True, False])
    
    df2.drop(['set_order'], axis=1, inplace=True)
    
    return df2

# In[circles_2]:
#run this analysis for all guilds
circle_raid_log = pd.DataFrame()

circle_raid_log = pd.concat([circle_raid_log, get_circle_data(us_source_raid_log)], axis = 0)
circle_raid_log = pd.concat([circle_raid_log, get_circle_data(bi_source_raid_log)], axis = 0)
circle_raid_log = pd.concat([circle_raid_log, get_circle_data(vn_source_raid_log)], axis = 0)
circle_raid_log = pd.concat([circle_raid_log, get_circle_data(ky_source_raid_log)], axis = 0)

#format the output
circle_raid_log_cols = circle_raid_log.columns[5:]
circle_raid_log[circle_raid_log_cols] = circle_raid_log[circle_raid_log_cols].apply(
    lambda c: pd.to_numeric(c, errors='coerce').round(0).astype('Int64')
)

# In[125]:


output_file = 'global_toplines.xlsx'

# Write all sheets
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    global_aggr_toplines.to_excel(writer, sheet_name='Global_agregated_toplines', index=False)
    meta_boss_df.to_excel(writer, sheet_name='Meta_boss_damage', index=False)
    boss_wise_points_pivot.to_excel(writer, sheet_name='Boss_wise_points', index=True)
    circle_raid_log.to_excel(writer, sheet_name='Circles_log', index=True)
    global_boss_df.to_excel(writer, sheet_name='Global_boss_data', index=False)
    global_detailed_toplines.to_excel(writer, sheet_name='Global_detailed_toplines', index=False)
    export_full_logs.to_excel(writer, sheet_name='Full_logs', index=False)

fixed_width = 20

# Load the workbook
wb = load_workbook(output_file)

# Loop over all sheets in the workbook

from openpyxl.utils import get_column_letter

for sheet in wb.worksheets:
    for i in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(i)
        sheet.column_dimensions[col_letter].width = fixed_width

# Save the workbook
wb.save(output_file)


# In[new_cell_2]:


# do colour and column width formatting for first tab
wb = load_workbook('global_toplines.xlsx')
ws = wb.active  # first sheet

# Find column index for "num_battles"
num_battles_col = None
for idx, cell in enumerate(ws[1], 1):  # header row
    if cell.value == 'num_battles':
        num_battles_col = idx
        break

if num_battles_col:
    col_letter = ws.cell(row=1, column=num_battles_col).column_letter
    # Apply 3-color gradient: red → yellow → green
    rule = ColorScaleRule(
        start_type='min', start_color='FF6347',    # red
        mid_type='percentile', mid_value=50, mid_color='FFFF00',  # yellow
        end_type='max', end_color='90EE90'         # green
    )
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

# Find column index for "total_points"
total_points_col = None
for idx, cell in enumerate(ws[1], 1):  # header row
    if cell.value == 'total_points':
        total_points_col = idx
        break

if total_points_col:
    col_letter = ws.cell(row=1, column=total_points_col).column_letter
    # Apply 3-color gradient: red → yellow → green
    rule = ColorScaleRule(
        start_type='min', start_color='FF6347',    # red
        mid_type='percentile', mid_value=50, mid_color='FFFF00',  # yellow
        end_type='max', end_color='90EE90'         # green
    )
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

# Tailor column width on the first tab
ws = wb.active 
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15


# Do row-wise colour coding in the 2nd tab
ws = wb['Meta_boss_damage']

# Define green fill
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Iterate over rows (skip header)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # Skip first column if non-numeric
    numeric_cells = row[1:]  
    max_value = max(cell.value for cell in numeric_cells if isinstance(cell.value, (int, float)))
    
    # Apply green fill to max value(s)
    for cell in numeric_cells:
        if cell.value == max_value:
            cell.fill = green_fill

ws.column_dimensions['A'].width = 21
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 25

# Delete first column in boss pivot df
ws = wb['Boss_wise_points']
ws.column_dimensions['A'].hidden = True  # 1-based index

# Delete first column in circles df
ws = wb['Circles_log']
ws.column_dimensions['A'].hidden = True  # 1-based index

# Save workbook
wb.save('global_toplines.xlsx')



# In[126]:


# Connect to Dropbox
# dbx is pulled from above

# Local file you want to upload
local_file = 'global_toplines' + '.xlsx'

# Path in Dropbox (inside your app folder or full Dropbox)
dropbox_path = "/" +'global_toplines' + '.xlsx'

# Upload the file
with open(local_file, "rb") as f:
    dbx.files_upload(
        f.read(), 
        dropbox_path, 
        mode=dropbox.files.WriteMode.overwrite)

print(f"File uploaded to Dropbox at: {dropbox_path}")




